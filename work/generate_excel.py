"""Generate Excel report from amoCRM link check results — ALL 104 deals."""

import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load data
with open("work/link_check_report.json", encoding="utf-8") as f:
    report = json.load(f)

with open("work/all_deals_full.json", encoding="utf-8") as f:
    all_deals = json.load(f)

wb = Workbook()

# ============================================================
# Color scheme
# ============================================================
FILL_OK = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")       # Green
FILL_EMPTY = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")     # Red/pink
FILL_FOLDER = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")    # Yellow
FILL_OTHER = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")     # Orange
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")    # Blue
FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
FONT_BOLD = Font(bold=True, size=11)
FONT_NORMAL = Font(size=10)
FONT_LINK = Font(size=10, color="0563C1", underline="single")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

STATUS_FILL = {
    "OK_FILE": FILL_OK,
    "EMPTY": FILL_EMPTY,
    "FOLDER": FILL_FOLDER,
    "OTHER_URL": FILL_OTHER,
    "GDRIVE_OTHER": FILL_OTHER,
}

STATUS_LABEL = {
    "OK_FILE": "Корректная ссылка",
    "EMPTY": "Пусто",
    "FOLDER": "Ссылка на папку",
    "OTHER_URL": "Мусор / не ссылка",
    "GDRIVE_OTHER": "GDrive (непонятно)",
}

SORT_ORDER = {"OTHER_URL": 0, "FOLDER": 1, "EMPTY": 2, "GDRIVE_OTHER": 3, "OK_FILE": 4}


def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


# ============================================================
# Sheet 1: Сводка
# ============================================================
ws_sum = wb.active
ws_sum.title = "Сводка"

ws_sum.merge_cells("A1:F1")
ws_sum.cell(1, 1, "Отчёт проверки ссылок — Воронка (amoCRM)").font = Font(bold=True, size=14)
ws_sum.cell(2, 1, f"Дата: {report['timestamp'][:10]}").font = FONT_NORMAL
ws_sum.cell(3, 1, f"Всего сделок проверено: {len(all_deals)}").font = FONT_NORMAL
ws_sum.cell(4, 1, f"Сделок с проблемами: {report['problems_count']}").font = Font(bold=True, size=11, color="FF0000")

# Pervichka summary table
row = 6
ws_sum.cell(row, 1, "Запись первички ссылка").font = FONT_BOLD
row += 1
for col, hdr in enumerate(["Статус", "Кол-во", "%"], 1):
    ws_sum.cell(row, col, hdr)
style_header(ws_sum, row, 3)

perv = report["summary"]["pervichka"]
total = len(all_deals)
for status in ["OK_FILE", "FOLDER", "EMPTY", "OTHER_URL", "GDRIVE_OTHER"]:
    cnt = perv.get(status, 0)
    if cnt == 0:
        continue
    row += 1
    ws_sum.cell(row, 1, STATUS_LABEL.get(status, status)).font = FONT_NORMAL
    ws_sum.cell(row, 2, cnt).font = FONT_NORMAL
    ws_sum.cell(row, 3, f"{cnt/total*100:.1f}%").font = FONT_NORMAL
    for c in range(1, 4):
        ws_sum.cell(row, c).fill = STATUS_FILL.get(status, PatternFill())
        ws_sum.cell(row, c).border = THIN_BORDER

# SS summary table
row += 2
ws_sum.cell(row, 1, "Запись СС ссылка").font = FONT_BOLD
row += 1
for col, hdr in enumerate(["Статус", "Кол-во", "%"], 1):
    ws_sum.cell(row, col, hdr)
style_header(ws_sum, row, 3)

ss = report["summary"]["ss"]
for status in ["OK_FILE", "FOLDER", "EMPTY", "OTHER_URL", "GDRIVE_OTHER"]:
    cnt = ss.get(status, 0)
    if cnt == 0:
        continue
    row += 1
    ws_sum.cell(row, 1, STATUS_LABEL.get(status, status)).font = FONT_NORMAL
    ws_sum.cell(row, 2, cnt).font = FONT_NORMAL
    ws_sum.cell(row, 3, f"{cnt/total*100:.1f}%").font = FONT_NORMAL
    for c in range(1, 4):
        ws_sum.cell(row, c).fill = STATUS_FILL.get(status, PatternFill())
        ws_sum.cell(row, c).border = THIN_BORDER

# Legend
row += 2
ws_sum.cell(row, 1, "Легенда цветов:").font = FONT_BOLD
for label, fill in [
    ("Корректная ссылка на файл", FILL_OK),
    ("Пусто (не заполнено)", FILL_EMPTY),
    ("Ссылка на папку (не на файл)", FILL_FOLDER),
    ("Мусор / не ссылка на GDrive", FILL_OTHER),
]:
    row += 1
    ws_sum.cell(row, 1, label).font = FONT_NORMAL
    ws_sum.cell(row, 1).fill = fill
    ws_sum.cell(row, 1).border = THIN_BORDER

for i, w in enumerate([35, 12, 8], 1):
    ws_sum.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# Sheet 2: Все сделки (ALL 104, sorted worst-first)
# ============================================================
ws_all = wb.create_sheet("Все сделки")

headers = [
    "ID сделки",
    "Имя клиента",
    "Этап воронки",
    "Первичка: статус",
    "Первичка: ссылка",
    "СС: статус",
    "СС: ссылка",
    "Ссылка на сделку",
]
for col, hdr in enumerate(headers, 1):
    ws_all.cell(1, col, hdr)
style_header(ws_all, 1, len(headers))
ws_all.auto_filter.ref = "A1:H1"
ws_all.freeze_panes = "A2"

# Sort: worst pervichka first, then worst SS
all_deals.sort(key=lambda d: (
    SORT_ORDER.get(d["pervichka_class"], 5),
    SORT_ORDER.get(d["ss_class"], 5),
    d["status"],
))

row = 1
for deal in all_deals:
    row += 1
    deal_id = deal["id"]
    deal_url = f"https://migratoramocrm.amocrm.ru/leads/detail/{deal_id}"

    ws_all.cell(row, 1, deal_id).font = FONT_NORMAL
    ws_all.cell(row, 2, deal["name"]).font = FONT_NORMAL
    ws_all.cell(row, 3, deal["status"]).font = FONT_NORMAL

    # Pervichka status + color
    perv_cls = deal["pervichka_class"]
    cell_ps = ws_all.cell(row, 4, STATUS_LABEL.get(perv_cls, perv_cls))
    cell_ps.font = FONT_NORMAL
    cell_ps.fill = STATUS_FILL.get(perv_cls, PatternFill())

    # Pervichka URL
    perv_url = deal.get("pervichka_url", "")
    cell_pu = ws_all.cell(row, 5, perv_url if perv_url else "—")
    if perv_url and perv_url.startswith("http") and len(perv_url) > 10:
        cell_pu.font = FONT_LINK
    else:
        cell_pu.font = FONT_NORMAL

    # SS status + color
    ss_cls = deal["ss_class"]
    is_expected_empty = deal["status"] == "FM Done" and ss_cls == "EMPTY"

    cell_ss = ws_all.cell(row, 6, STATUS_LABEL.get(ss_cls, ss_cls))
    cell_ss.font = FONT_NORMAL
    if is_expected_empty:
        # FM Done + empty SS is expected, use light gray
        cell_ss.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    else:
        cell_ss.fill = STATUS_FILL.get(ss_cls, PatternFill())

    # SS URL
    ss_url = deal.get("ss_url", "")
    cell_su = ws_all.cell(row, 7, ss_url if ss_url else "—")
    if ss_url and ss_url.startswith("http") and len(ss_url) > 10:
        cell_su.font = FONT_LINK
    else:
        cell_su.font = FONT_NORMAL

    # Link to deal
    cell_link = ws_all.cell(row, 8, deal_url)
    cell_link.font = FONT_LINK
    cell_link.hyperlink = deal_url

    # Borders
    for c in range(1, len(headers) + 1):
        ws_all.cell(row, c).border = THIN_BORDER
        ws_all.cell(row, c).alignment = Alignment(vertical="center", wrap_text=False)

# Column widths
widths = [12, 42, 24, 22, 60, 22, 60, 48]
for i, w in enumerate(widths, 1):
    ws_all.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# Sheet 3: По типам проблем
# ============================================================
ws_type = wb.create_sheet("По типам проблем")

# Only problem deals
problem_deals = [d for d in all_deals if d["pervichka_class"] != "OK_FILE" or (
    d["ss_class"] != "OK_FILE" and not (d["status"] == "FM Done" and d["ss_class"] == "EMPTY")
)]

by_type = {}
for p in problem_deals:
    types = []
    if p["pervichka_class"] != "OK_FILE":
        types.append(f"Первичка: {STATUS_LABEL.get(p['pervichka_class'], p['pervichka_class'])}")
    ss_cls = p["ss_class"]
    is_expected = p["status"] == "FM Done" and ss_cls == "EMPTY"
    if ss_cls != "OK_FILE" and not is_expected:
        types.append(f"СС: {STATUS_LABEL.get(ss_cls, ss_cls)}")
    key = " + ".join(types) if types else "Нет проблем"
    by_type.setdefault(key, []).append(p)

row = 0
for ptype, deals in sorted(by_type.items(), key=lambda x: -len(x[1])):
    row += 1
    ws_type.cell(row, 1, f"{ptype} ({len(deals)} сделок)").font = Font(bold=True, size=12)
    ws_type.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

    row += 1
    for col, hdr in enumerate(["ID", "Клиент", "Этап", "Первичка URL", "СС URL"], 1):
        ws_type.cell(row, col, hdr)
    style_header(ws_type, row, 5)

    for d in deals:
        row += 1
        ws_type.cell(row, 1, d["id"]).font = FONT_NORMAL
        ws_type.cell(row, 2, d["name"]).font = FONT_NORMAL
        ws_type.cell(row, 3, d["status"]).font = FONT_NORMAL
        ws_type.cell(row, 4, d.get("pervichka_url", "") or "—").font = FONT_NORMAL
        ws_type.cell(row, 5, d.get("ss_url", "") or "—").font = FONT_NORMAL
        for c in range(1, 6):
            ws_type.cell(row, c).border = THIN_BORDER

    row += 1  # gap

for i, w in enumerate([12, 42, 24, 60, 60], 1):
    ws_type.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Save
# ============================================================
output = "work/amocrm_link_check_report.xlsx"
wb.save(output)
print(f"Excel saved: {output}")
print(f"Sheets: Сводка, Все сделки ({len(all_deals)} rows), По типам проблем ({len(problem_deals)} rows)")
